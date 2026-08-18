import os
import json
import urllib.parse
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import google.generativeai as genai
from .config import settings
from .database import add_scan_log, track_llm_call
from .schemas import ScopeContext
from .mcp_explorer import derive_features_from_elements
from typing import Optional

# Per-feature-type templates for the structural mock generator, keyed to the same feature
# types produced by derive_features_from_elements() so the fallback path (which runs whenever
# Gemini is unavailable or a stage fails) still covers every discovered element category, not
# just forms/buttons.
MOCK_TYPE_TEMPLATES = {
    "Create": {"priority": "High", "scenario": "Verify submitting {feat} with structural inputs",
                "steps": "1. Locate the form ({feat}).\n2. Populate fields: {fields}.\n3. Click the submit control.",
                "expected": "The form is submitted successfully.", "rule": "Form fields must accept valid formats."},
    "Action": {"priority": "Medium", "scenario": "Verify triggering {feat}",
               "steps": "1. Locate the control for '{feat}'.\n2. Trigger the action.",
               "expected": "The '{feat}' action responds correctly.", "rule": "Interactions must respond correctly."},
    "Read": {"priority": "High", "scenario": "Verify {feat} loads and displays discovered data",
             "steps": "1. Navigate to page: {url}.\n2. Verify {feat} renders with data.",
             "expected": "{feat} is visible and populated.", "rule": "Listing data must load without error."},
    "Search": {"priority": "Medium", "scenario": "Verify {feat} narrows the listing to matching results",
               "steps": "1. Enter a known value into the search control.\n2. Confirm results update to match.",
               "expected": "Only matching rows remain visible.", "rule": "Search must narrow results to relevant matches."},
    "Filter": {"priority": "Medium", "scenario": "Verify {feat} narrows the listing to matching records",
               "steps": "1. Trigger the {feat} control.\n2. Confirm the listing updates.",
               "expected": "Only records matching the filter remain visible.", "rule": "Filters must accurately narrow the dataset."},
    "Sort": {"priority": "Low", "scenario": "Verify {feat} reorders the listing",
             "steps": "1. Trigger the {feat} control.\n2. Confirm row order changes accordingly.",
             "expected": "Rows are reordered according to the selected sort.", "rule": "Sorting must correctly reorder displayed data."},
    "Pagination": {"priority": "Medium", "scenario": "Verify {feat} navigates between result pages",
                   "steps": "1. Trigger the pagination control.\n2. Confirm a different page of results loads.",
                   "expected": "A different set of rows is displayed.", "rule": "Pagination must load distinct result pages."},
    "Toggle": {"priority": "Low", "scenario": "Verify {feat} can be toggled",
               "steps": "1. Locate the checkbox for '{feat}'.\n2. Toggle it on and off.",
               "expected": "The checkbox state changes and persists as expected.", "rule": "Checkbox state changes must be reflected in the UI."},
    "Navigation": {"priority": "Low", "scenario": "Verify {feat} switches the active view",
                   "steps": "1. Trigger the control: {feat}.\n2. Confirm the corresponding view is displayed.",
                   "expected": "The selected view's content is displayed.", "rule": "Navigation controls must display the correct associated content."},
    "Select": {"priority": "Low", "scenario": "Verify {feat} allows selecting a discovered value",
               "steps": "1. Open the control for '{feat}'.\n2. Choose an available option.",
               "expected": "The selected value is applied.", "rule": "Selection controls must accept a valid discovered option."},
    "Upload": {"priority": "Low", "scenario": "Verify {feat} is present and available",
               "steps": "1. Locate the control for '{feat}'.",
               "expected": "The '{feat}' control is available for use.", "rule": "File transfer controls must be present and accessible."},
    "Download": {"priority": "Low", "scenario": "Verify {feat} is present and available",
                 "steps": "1. Locate the control for '{feat}'.",
                 "expected": "The '{feat}' control is available for use.", "rule": "File transfer controls must be present and accessible."},
    "Modal": {"priority": "Low", "scenario": "Verify the {feat} dialog is displayed with the expected content",
              "steps": "1. Trigger the action associated with '{feat}'.\n2. Confirm the dialog opens and displays its content.",
              "expected": "The '{feat}' dialog is displayed correctly.", "rule": "Dialogs must present their content correctly when opened."},
    "Drawer": {"priority": "Low", "scenario": "Verify the {feat} panel is displayed with the expected content",
               "steps": "1. Trigger the action associated with '{feat}'.\n2. Confirm the panel opens and displays its content.",
               "expected": "The '{feat}' panel is displayed correctly.", "rule": "Panels must present their content correctly when opened."},
}

# Phrases that indicate a scenario describes a raw DOM/CSS implementation detail rather than a
# user-facing action or observable UI behavior. Deliberately generic (no app-specific terms) so
# it works as a safety net on any target - applied to both the mock-fallback path (which has no
# LLM to instruct) and as a post-Quality-Gate pass on the real pipeline (defense in depth, since
# an LLM could still invent similar phrasing on its own).
_STRUCTURAL_NOISE_PHRASES = [
    "exists in the page structure", "exists in the dom", "element exists",
    "div exists", "class exists", "css class", "dom node exists", "selector exists",
]

def is_structural_noise_scenario(scenario: str) -> bool:
    """True if a scenario describes DOM/CSS existence rather than a user-facing behavior."""
    s = (scenario or "").lower()
    return any(phrase in s for phrase in _STRUCTURAL_NOISE_PHRASES)

def log_pipeline(scan_id: str, scope_ctx: Optional[ScopeContext], tag: str, message: str):
    scope = "unknown"
    parent = "None"
    selected = "None"
    if scope_ctx:
        scope = scope_ctx.scan_scope or "unknown"
        parent = scope_ctx.parent_module or "None"
        selected = scope_ctx.selected_module or "None"
    prefix = f"[{tag}] [scan_id={scan_id}] [scope={scope}] [parent={parent}] [selected={selected}]"
    add_scan_log(scan_id, f"{prefix} {message}")

def get_mock_test_cases(app_map: dict = None, scope_ctx: ScopeContext = None) -> list:
    """
    Generates dynamic structural mock test cases conforming to the final schema when Gemini is not configured.
    It derives test cases dynamically from the app_map nodes matching the selected scope, ensuring no hardcoded EMS leaks.
    """
    nodes = app_map.get("nodes", []) if app_map else []
    
    # Filter nodes by scope
    if scope_ctx and scope_ctx.scan_scope in ("module", "MODULE", "MODULE_WISE"):
        target_paths = [urllib.parse.urlparse(u).path.lower().strip('/') for u in scope_ctx.target_urls]
        filtered_nodes = []
        for n in nodes:
            url_path = urllib.parse.urlparse(n.get("url", "")).path.lower().strip('/')
            if any(tp in url_path for tp in target_paths) if target_paths else False:
                filtered_nodes.append(n)
        nodes = filtered_nodes

    if not nodes:
        return []

    mock_suite = []
    tc_counter = 1
    
    parent_name = scope_ctx.parent_module if (scope_ctx and scope_ctx.parent_module) else "General"
    selected_name = scope_ctx.selected_module if (scope_ctx and scope_ctx.selected_module) else "General"

    for node in nodes:
        page_label = node.get("label", "Target Page")
        url = node.get("url", "")
        elements = node.get("elements", {})
        
        # 1. Base Exploration Test Case
        mock_suite.append({
            "module": parent_name,
            "submodule": selected_name,
            "page": page_label,
            "feature": "Explore main view",
            "test_case_id": f"TC-{selected_name.upper()[:6]}-{tc_counter:03d}",
            "scenario": f"Verify user can load and view the {page_label} interface",
            "test_type": "Smoke",
            "priority": "High",
            "preconditions": f"Navigate to page: {url}",
            "test_data": "None",
            "test_steps": f"1. Navigate to URL: {url}.\n2. Wait for page load.\n3. Verify the main dashboard/listing loaded.",
            "expected_result": f"The page {page_label} is displayed and elements are visible.",
            "business_rule": "The interface must render without errors.",
            "dependencies": "Target application running.",
            "source_classification": "STRUCTURAL_EVIDENCE",
            "evidence_reference": f"EVID-STRUCT-{tc_counter:03d}",
            "observed_behavior": "STRUCTURAL_EVIDENCE",
            "automation_candidate": "Yes",
            "automation_selector": "",
            "execution_engine": "NONE"
        })
        tc_counter += 1

        # 2. One test case per discovered feature, across every element category the crawler
        # extracted (forms, buttons, tables, search/filter/sort/pagination controls,
        # checkboxes, radios, tabs, selects, uploads/downloads, modals/drawers) - not just
        # forms and buttons - so the fallback path stays proportional to module complexity.
        page_features, _ = derive_features_from_elements(elements)
        for feat in page_features:
            # Tabs may be navigation into a sibling module rather than an in-page view switch -
            # scope-filter those by href the same way pages themselves are filtered above.
            feat_href = feat.get("href")
            if feat_href and scope_ctx and scope_ctx.scan_scope in ("module", "MODULE", "MODULE_WISE") and scope_ctx.target_urls:
                href_target_paths = [urllib.parse.urlparse(u).path.lower().strip('/') for u in scope_ctx.target_urls]
                feat_href_path = urllib.parse.urlparse(feat_href).path.lower().strip('/')
                is_feat_dashboard = any(x in feat_href_path for x in ["dashboard", "login", "logout"])
                is_feat_target = any(tp in feat_href_path for tp in href_target_paths)
                if not is_feat_target and not is_feat_dashboard:
                    continue
            f_type = feat["feature_type"]
            f_name = feat["feature_name"]
            fields_desc = ", ".join(f.get("name", "field") for f in feat.get("fields", []))
            tmpl = MOCK_TYPE_TEMPLATES.get(f_type, MOCK_TYPE_TEMPLATES["Action"])

            mock_suite.append({
                "module": parent_name,
                "submodule": selected_name,
                "page": page_label,
                "feature": f_name,
                "test_case_id": f"TC-{selected_name.upper()[:6]}-{tc_counter:03d}",
                "scenario": tmpl["scenario"].format(feat=f_name, url=url),
                "test_type": "Functional",
                "priority": tmpl["priority"],
                "preconditions": f"Open page: {url}",
                "test_data": fields_desc or "None",
                "test_steps": tmpl["steps"].format(feat=f_name, url=url, fields=fields_desc or "discovered fields"),
                "expected_result": tmpl["expected"].format(feat=f_name, url=url),
                "business_rule": tmpl["rule"],
                "dependencies": f"{f_name} present in DOM.",
                "source_classification": "STRUCTURAL_EVIDENCE",
                "evidence_reference": f"EVID-STRUCT-{tc_counter:03d}",
                "observed_behavior": "STRUCTURAL_EVIDENCE",
                "automation_candidate": "Yes",
                "automation_selector": feat.get("locator_hint", ""),
                "execution_engine": "NONE"
            })
            tc_counter += 1

    # No LLM runs in this fallback path, so apply the deterministic structural-noise guard
    # directly rather than relying on the (LLM-only) Quality Gate.
    return [tc for tc in mock_suite if not is_structural_noise_scenario(tc.get("scenario", ""))]

def query_llm(model, prompt: str, scan_id: str, stage_name: str) -> str:
    """Helper to query the configured generative model with logs."""
    add_scan_log(scan_id, f"[GENERATOR] [{stage_name}] Dispatching reasoning task to Gemini...")
    try:
        with track_llm_call(scan_id, stage_name):
            response = model.generate_content(prompt)
        text = response.text.strip()
        
        if "```json" in text:
            text = text.split("```json")[1].split("```")[0]
        elif "```" in text:
            text = text.split("```")[1].split("```")[0]
        return text.strip()
    except Exception as e:
        add_scan_log(scan_id, f"[WARNING] [{stage_name}] LLM dispatch failed: {e}")
        raise e

def run_quality_gate(model, test_cases: list, scan_id: str, scope_ctx: ScopeContext = None) -> list:
    """Quality Gate checking: Rejects/rewrites generic test cases into highly specific ones based on observed data."""
    if not test_cases:
        return []
        
    log_pipeline(scan_id, scope_ctx, "QUALITY-GATE", "Starting Final Quality Gate Audit...")
    cleaned_test_cases = []
    batch_size = 8
    
    scan_scope_str = scope_ctx.scan_scope if scope_ctx else "entire"
    parent_mod = scope_ctx.parent_module if scope_ctx else ""
    selected_mod = scope_ctx.selected_module if scope_ctx else ""
    
    for i in range(0, len(test_cases), batch_size):
        batch = test_cases[i:i+batch_size]
        
        prompt = f"""
        You are a Senior QA Automation Architect acting as a Quality Gate.
        Review the following batch of generated test cases.
        Scan Scope: {scan_scope_str}
        Target Parent Module: {parent_mod}
        Target Selected Module: {selected_mod}
        
        BATCH:
        {json.dumps(batch, indent=2)}
        
        Rules:
        1. REWRITE any generic test cases. Reject generic descriptions like "Verify CRUD functionality", "Verify Add User", "Verify search", or "Verify validation".
        2. Rewrite them to be highly specific, referencing the discovered elements (selectors, input names, labels) and exact observed validation behaviors.
        3. REJECT any scenario that describes a raw DOM/CSS/implementation detail instead of a user-facing action or observable UI behavior. A scenario must answer "what does the user do?" and "what should the user see?" - never "what DOM element/CSS class exists?".
           Bad (reject): "Verify input#email exists.", "Verify modal-box exists.", "Verify drawer-side exists.", "Verify class xyz exists.", "Verify div exists.", "Verify selector exists."
           Good (accept): "Verify the Email field is displayed in the Add form.", "Verify clicking Add opens the creation dialog.", "Verify entering invalid data displays the validation message.", "Verify applying the Status filter updates the displayed records."
           If a scenario names a UI container (dialog, drawer, panel) but only asserts its existence, either rewrite it around the actual workflow that opens it (if evidence shows one) or reject it - never keep an "exists in the page/DOM structure" style scenario.
        4. For Module Wise scans (scan_scope = module), if any test case belongs to a different module than Target Selected Module ({selected_mod}), REJECT it (do not include it in the output list).
           Reject generic scenarios. They must be page/feature/field-specific.
           For example:
           - If a 'Company' field is observed: "Verify submitting the Add Client form without entering a value in the Company field displays the observed required-field validation."
           - If an 'Email' field is observed: "Verify entering an invalid email format in the Email field triggers the observed validation message."
           Do NOT invent the exact validation message unless it was actually observed.
        5. If a scenario has source_classification or observed_behavior equal to "SKIPPED_SCENARIO" (e.g. because it was blocked by the safety gateway, like Delete):
           - Rewrite it to focus on safe behavior, e.g. "Verify the Delete action for a client opens a confirmation dialog."
           - Do NOT write "Verify client is successfully deleted" because deletion was never executed.
        6. Drop duplicate scenarios.
        7. Do NOT modify the classification values like 'observed_behavior' or 'source_classification'. Keep them unchanged.
        8. Return ONLY a JSON list matching the input structure exactly.
        """
        try:
            res_txt = query_llm(model, prompt, scan_id, "QUALITY GATE")
            cleaned = json.loads(res_txt)
            cleaned_test_cases.extend(cleaned)
        except Exception as e:
            log_pipeline(scan_id, scope_ctx, "QUALITY-GATE", f"[WARNING] Batch parsing failed: {e}. Preserving original batch cases.")
            cleaned_test_cases.extend(batch)
            
    return cleaned_test_cases

def generate_test_cases(scan_id: str, app_map: dict, user_description: str, api_key: str = None, 
                        model_name: str = None, mcp_behavior: str = None, scope_ctx: ScopeContext = None) -> list:
    """Invokes the multi-stage E2E QA reasoning pipeline using Gemini and structured MCP behavioral data."""
    log_pipeline(scan_id, scope_ctx, "GENERATOR", "Ingesting Structured Behavioral Discovery Data...")

    scan_scope_str = scope_ctx.scan_scope if scope_ctx else "entire"

    gen_key = api_key or settings.GEMINI_API_KEY
    if not gen_key:
        log_pipeline(scan_id, scope_ctx, "GENERATOR", "[WARNING] No GEMINI_API_KEY provided. Generating mock test cases...")
        mock_cases = get_mock_test_cases(app_map, scope_ctx)
        
        total_modules = len(set(tc.get("module", "") for tc in mock_cases))
        total_pages = len(set(tc.get("page", "") for tc in mock_cases))
        total_features = len(set((tc.get("module", ""), tc.get("feature", "")) for tc in mock_cases))
        
        observed_count = sum(1 for tc in mock_cases if tc.get("observed_behavior") == "OBSERVED_BEHAVIOR")
        structural_count = sum(1 for tc in mock_cases if tc.get("observed_behavior") == "STRUCTURAL_EVIDENCE")
        inferred_count = sum(1 for tc in mock_cases if tc.get("observed_behavior") == "INFERRED_SCENARIO")
        skipped_count = sum(1 for tc in mock_cases if tc.get("observed_behavior") == "SKIPPED_SCENARIO")
        
        coverage_pct = int((observed_count / max(1, observed_count + structural_count + inferred_count + skipped_count)) * 100)
        
        add_scan_log(scan_id, "==================================================")
        add_scan_log(scan_id, "               VALIDATION METRICS                 ")
        add_scan_log(scan_id, "==================================================")
        add_scan_log(scan_id, f" - Total Modules: {total_modules}")
        add_scan_log(scan_id, f" - Total Pages: {total_pages}")
        add_scan_log(scan_id, f" - Total Features: {total_features}")
        add_scan_log(scan_id, f" - Observed Behaviors: {observed_count}")
        add_scan_log(scan_id, f" - Structurally Discovered Behaviors: {structural_count}")
        add_scan_log(scan_id, f" - Inferred Behaviors: {inferred_count}")
        add_scan_log(scan_id, f" - Skipped Behaviors: {skipped_count}")
        add_scan_log(scan_id, f" - Coverage %: {coverage_pct}%")
        add_scan_log(scan_id, f" - Generated Test Cases: {len(mock_cases)}")
        add_scan_log(scan_id, "==================================================")
        
        # Display Feature Coverage Matrix for mock cases
        add_scan_log(scan_id, "=======================================================================================")
        add_scan_log(scan_id, "                           FEATURE COVERAGE MATRIX                                     ")
        add_scan_log(scan_id, "=======================================================================================")
        add_scan_log(scan_id, " Module | Feature | Observed | Structural | Inferred | Skipped | Coverage % ")
        add_scan_log(scan_id, "---------------------------------------------------------------------------------------")
        unique_features = set((tc.get("module", ""), tc.get("feature", "")) for tc in mock_cases)
        for mod, feat in sorted(unique_features):
            feat_cases = [tc for tc in mock_cases if tc.get("module") == mod and tc.get("feature") == feat]
            obs = sum(1 for tc in feat_cases if tc.get("observed_behavior") == "OBSERVED_BEHAVIOR")
            struct = sum(1 for tc in feat_cases if tc.get("observed_behavior") == "STRUCTURAL_EVIDENCE")
            inf = sum(1 for tc in feat_cases if tc.get("observed_behavior") == "INFERRED_SCENARIO")
            skip = sum(1 for tc in feat_cases if tc.get("observed_behavior") == "SKIPPED_SCENARIO")
            total = obs + struct + inf + skip
            feat_pct = int((obs / max(1, total)) * 100)
            add_scan_log(scan_id, f" {mod} | {feat} | {obs} | {struct} | {inf} | {skip} | {feat_pct}% ")
        add_scan_log(scan_id, "=======================================================================================")
        
        return mock_cases

    # Ingest structured behavior modules
    mcp_behavior_data = {}
    safety_log = []
    coverage_gaps = []
    feature_registry = {} # Python telemetry override source mapping
    locator_registry = {} # Automation metadata (best-effort selectors), kept separate from scenario text
    engine_registry = {} # execution_engine (MCP / PLAYWRIGHT_CORE / NONE) per (module, feature)
    total_mcp_actions = 0

    if mcp_behavior:
        try:
            behavior_json = json.loads(mcp_behavior)
            safety_log = behavior_json.get("safety_actions_log", [])
            coverage_gaps = behavior_json.get("coverage_gaps", [])
            # Real behavioral evidence can come from either engine - counting only MCP actions
            # here would incorrectly downgrade legitimate Playwright Core OBSERVED_BEHAVIOR
            # results to INFERRED_SCENARIO whenever the fallback engine (not MCP) did the work.
            behavior_metrics = behavior_json.get("metrics", {})
            total_mcp_actions = behavior_metrics.get("mcp_actions_count", 0) + behavior_metrics.get("pw_core_successful_calls", 0)

            if total_mcp_actions == 0:
                add_scan_log(scan_id, "[GENERATOR] Enforcing Behavioral Evidence Rule: 0 behavioral actions detected (MCP or Playwright Core). Disabling OBSERVED_BEHAVIOR.")

            for mod in behavior_json.get("modules", []):
                mod_name = mod["module_name"]
                mcp_behavior_data[mod_name] = mod
                # Register all features from Pass 2 behavior mappings
                for feat in mod.get("features", []):
                    f_name = feat.get("feature_name", feat.get("name", "Explore main view"))
                    raw_status = feat.get("source_classification", "STRUCTURAL_EVIDENCE")
                    # Map to exact final classifications
                    if total_mcp_actions == 0:
                        status = "STRUCTURAL_EVIDENCE" if raw_status == "STRUCTURAL_EVIDENCE" else "INFERRED_SCENARIO"
                    else:
                        if raw_status == "OBSERVED_BEHAVIOR" or raw_status == "OBSERVED":
                            status = "OBSERVED_BEHAVIOR"
                        elif raw_status == "SKIPPED_SCENARIO" or raw_status == "SKIPPED":
                            status = "SKIPPED_SCENARIO"
                        elif raw_status == "INFERRED_SCENARIO" or raw_status == "INFERRED":
                            status = "INFERRED_SCENARIO"
                        else:
                            status = "STRUCTURAL_EVIDENCE"
                    feature_registry[(mod_name.strip().lower(), f_name.strip().lower())] = status
                    locator_hint = feat.get("locator_hint")
                    if locator_hint:
                        locator_registry[(mod_name.strip().lower(), f_name.strip().lower())] = locator_hint
                    engine_registry[(mod_name.strip().lower(), f_name.strip().lower())] = feat.get("execution_engine", "NONE")

            add_scan_log(scan_id, f"[GENERATOR] Compiled Python Telemetry Registry with {len(feature_registry)} feature keys.")
        except Exception as e:
            add_scan_log(scan_id, f"[WARNING] Failed to parse structured behavior telemetry: {e}")

    try:
        genai.configure(api_key=gen_key)
        selected_model = model_name or "gemini-1.5-flash"
        model = genai.GenerativeModel(selected_model)
        
        crawled_nodes = app_map.get("nodes", [])
        crawled_edges = app_map.get("edges", [])

        # For MODULE scope, constrain the LLM's input up front to only the resolved target
        # page(s) instead of handing it the entire crawl and relying solely on the final
        # substring filter (STAGE 8) to clean up afterwards - that filter is a safety net, not
        # the only enforcement. Prerequisite pages (e.g. authentication) are summarized as a
        # precondition note rather than passed as full page content, so the LLM has no raw
        # material to invent unrelated Login/Settings/etc. test cases from.
        is_module_scope = scope_ctx and scope_ctx.scan_scope in ("module", "MODULE", "MODULE_WISE")
        scope_precondition_note = ""
        if is_module_scope and scope_ctx.target_urls:
            target_paths = [urllib.parse.urlparse(u).path.lower().strip('/') for u in scope_ctx.target_urls]
            in_scope_nodes = [
                n for n in crawled_nodes
                if any(tp in urllib.parse.urlparse(n.get("url", "")).path.lower().strip('/') for tp in target_paths)
            ]
            if in_scope_nodes:
                out_of_scope_count = len(crawled_nodes) - len(in_scope_nodes)
                if out_of_scope_count:
                    scope_precondition_note = (
                        f"Precondition: reaching {scope_ctx.parent_module} -> {scope_ctx.selected_module} required "
                        f"navigating through {out_of_scope_count} other prerequisite page(s) (e.g. authentication). "
                        "Do NOT generate modules or test cases for those prerequisite pages themselves."
                    )
                in_scope_urls = {n.get("url", "") for n in in_scope_nodes}
                crawled_edges = [e for e in crawled_edges if e.get("source") in in_scope_urls and e.get("target") in in_scope_urls]
                crawled_nodes = in_scope_nodes

        # --- STAGE 1: Build Application Map & Module Inventory ---
        add_scan_log(scan_id, "[GENERATOR] [STAGE 1] Creating Logical Application Map & Module Inventories...")
        stage1_prompt = f"""
        Analyze the following crawled web application details and construct a logical, hierarchical Application Map and Module Inventory.

        CRAWLED PAGES DETAILS:
        {json.dumps(crawled_nodes, indent=2)}

        CRAWLED TRANSITIONS:
        {json.dumps(crawled_edges, indent=2)}

        STRUCTURED BEHAVIORAL MAP:
        {json.dumps(list(mcp_behavior_data.values()), indent=2)}

        SCAN SCOPE CONFIGURATION:
        Scope: {scan_scope_str}
        {f"Target module: {scope_ctx.parent_module} -> {scope_ctx.selected_module}. Only build modules from the CRAWLED PAGES DETAILS above - do not invent additional modules outside that list." if is_module_scope else ""}
        {scope_precondition_note}
        User description: "{user_description or 'Focus on dynamic form validations.'}"

        Return ONLY a JSON object matching the schema.
        """

        stage1_res = query_llm(model, stage1_prompt, scan_id, "STAGE 1: Map & Inventory")
        try:
            inventory_data = json.loads(stage1_res)
        except Exception as e:
            add_scan_log(scan_id, f"[WARNING] Failed to parse Stage 1 Map JSON: {e}.")
            inventory_data = {"application_map": {"modules": []}, "module_inventory": []}
            for node in crawled_nodes:
                url_path = urllib.parse.urlparse(node["url"]).path or "/"
                if is_module_scope:
                    mod_name = f"{scope_ctx.parent_module} - {scope_ctx.selected_module}"
                else:
                    mod_name = node.get("label") or url_path
                node_features, _ = derive_features_from_elements(node.get("elements", {}))
                feature_names = [f["feature_name"] for f in node_features] or ["General Functionality"]
                inventory_data["module_inventory"].append({
                    "module_name": mod_name,
                    "features": feature_names,
                    "fields": [],
                    "states": [],
                    "dependencies": []
                })
        
        # --- STAGE 2: Behavior Analysis & Scenario Matrices ---
        add_scan_log(scan_id, "[GENERATOR] [STAGE 2] Performing deep Behavior Analysis & Scenario Matrices...")
        all_scenario_matrices = []
        for module in inventory_data.get("module_inventory", []):
            # Stage 1's module_inventory is free-form Gemini JSON, not schema-validated - don't
            # trust it to always use the exact key name we asked for.
            module_name = module.get("module_name") or module.get("name") or "Unnamed Module"
            module_pages_elements = []
            for node in crawled_nodes:
                module_pages_elements.append({
                    "page_name": node.get("label"),
                    "url": node["url"],
                    "elements": node.get("elements", {})
                })

            behavior_obs = mcp_behavior_data.get(module_name, {})

            stage2_prompt = f"""
            Perform a detailed Behavior Analysis and construct a Scenario Matrix for the following module.

            MODULE NAME: {module_name}
            DISCOVERED FEATURES: {json.dumps(module.get("features", []))}

            ELEMENTS IN THE MODULE PAGES:
            {json.dumps(module_pages_elements[:10], indent=2)}

            STRUCTURED BEHAVIOR OBSERVATIONS (PLAYWRIGHT MCP):
            {json.dumps(behavior_obs, indent=2)}

            SAFETY SKIPPED ACTIONS LOG:
            {json.dumps(safety_log, indent=2)}

            For EACH discovered feature, consider whether each of the following behavior
            categories applies, and only emit a scenario for a category when the ELEMENTS or
            STRUCTURED BEHAVIOR OBSERVATIONS above actually provide evidence it applies to that
            specific feature - do not generate a category just because it exists in this list:
            - happy path (successful use with valid input)
            - required-field validation (only if a field has required=true or a required
              indicator was observed)
            - invalid input / format validation (only if the field has a type like email/number
              that implies a format, or a validation message was observed)
            - boundary values (only for numeric/length-constrained fields)
            - search / filtering / sorting / pagination (only if that control was discovered)
            - create / update / delete (only if a create form or an edit/delete control was
              discovered - reference the SAFETY SKIPPED ACTIONS LOG for delete, since it is
              never actually executed)
            - cancel / confirmation dialogs (only if a modal/dialog was discovered near the
              action)
            - error handling / success handling (only if a validation message, toast, or
              transition was observed)
            - persistence / state changes after navigating away and back (only if a transition
              was observed)
            - navigation between pages or tabs (only if discovered nav links or tabs exist)

            Each scenario must reference the actual discovered field names, labels, button text,
            or table headers from ELEMENTS/STRUCTURED BEHAVIOR OBSERVATIONS above - never a
            generic phrase like "Verify CRUD functionality" or "Verify the interface loads".

            Return ONLY a JSON array of features containing scenarios.
            """
            try:
                stage2_res = query_llm(model, stage2_prompt, scan_id, f"STAGE 2: Behavior - {module_name}")
                module_scenarios = json.loads(stage2_res)
                all_scenario_matrices.append({
                    "module_name": module_name,
                    "features": module_scenarios
                })
            except Exception as e:
                add_scan_log(scan_id, f"[WARNING] Failed behavior analysis for {module_name}: {e}.")

        # --- STAGE 3: Cross-Module E2E Analysis ---
        add_scan_log(scan_id, "[GENERATOR] [STAGE 3] Performing Cross-Module E2E Workflow Analysis...")
        stage3_prompt = f"""
        Analyze modules and navigation connections, then map logical cross-module End-to-End (E2E) workflows.
        
        MODULE INVENTORIES:
        {json.dumps(inventory_data.get("module_inventory", []), indent=2)}
        
        STRUCTURED BEHAVIORAL MAP:
        {json.dumps(list(mcp_behavior_data.values()), indent=2)}
        
        Return ONLY a JSON list of workflows.
        """
        try:
            stage3_res = query_llm(model, stage3_prompt, scan_id, "STAGE 3: E2E Workflows")
            e2e_workflows = json.loads(stage3_res)
        except Exception as e:
            e2e_workflows = []

        # --- STAGE 4: Test Case Generation ---
        add_scan_log(scan_id, "[GENERATOR] [STAGE 4] Generating detailed Test Case suite...")
        generated_test_cases = []
        tc_count = 0
        
        for matrix in all_scenario_matrices:
            m_name = matrix["module_name"]
            for feature in matrix["features"]:
                f_name = feature.get("feature_name", feature.get("name", "Explore main view"))
                scenarios = feature.get("scenarios", [])
                if not scenarios:
                    continue
                
                batch_size = 5
                for i in range(0, len(scenarios), batch_size):
                    batch = scenarios[i:i+batch_size]
                    stage4_prompt = f"""
                    Generate detailed manual test cases.
                    MODULE: {m_name}
                    FEATURE: {f_name}
                    SCENARIOS: {json.dumps(batch, indent=2)}
                    Return ONLY JSON list of test cases.
                    """
                    try:
                        stage4_res = query_llm(model, stage4_prompt, scan_id, f"STAGE 4: Test Cases - {f_name}")
                        cases = json.loads(stage4_res)
                        for c in cases:
                            tc_count += 1
                            c["test_case_id"] = f"TC-{m_name.upper().replace(' ', '_')[:6]}-{tc_count}"
                            generated_test_cases.append(c)
                    except Exception as e:
                        add_scan_log(scan_id, f"[WARNING] Batch test generation failed: {e}")

        # E2E Workflows test cases
        if e2e_workflows:
            for workflow in e2e_workflows:
                wf_name = workflow.get("workflow_name") or workflow.get("name") or "Workflow"
                stage4_e2e_prompt = f"""
                Generate E2E case for: {wf_name}
                Return ONLY JSON.
                """
                try:
                    stage4_e2e_res = query_llm(model, stage4_e2e_prompt, scan_id, f"STAGE 4: E2E - {wf_name}")
                    e2e_cases = json.loads(stage4_e2e_res)
                    for ec in e2e_cases:
                        tc_count += 1
                        ec["test_case_id"] = f"TC-E2E-{tc_count}"
                        generated_test_cases.append(ec)
                except Exception as e:
                    add_scan_log(scan_id, f"[WARNING] E2E test case generation failed: {e}")

        # --- STAGE 5: Coverage Auditor & Gap Filler ---
        add_scan_log(scan_id, "[GENERATOR] [STAGE 5] Launching Test Coverage Auditor...")
        all_features = []
        for m in inventory_data.get("module_inventory", []):
            m_name = m.get("module_name") or m.get("name") or "Unnamed Module"
            for f in m.get("features", []):
                all_features.append({"module": m_name, "feature": f})
        
        covered = set()
        for tc in generated_test_cases:
            covered.add((tc.get("module", "").strip().lower(), tc.get("feature", "").strip().lower()))
            
        gaps = [f for f in all_features if (f["module"].strip().lower(), f["feature"].strip().lower()) not in covered]
        for gap in coverage_gaps:
            g_mod = gap.get("module")
            g_feat = gap.get("feature")
            if g_mod and g_feat:
                if (g_mod.strip().lower(), g_feat.strip().lower()) not in covered:
                    gaps.append({"module": g_mod, "feature": g_feat})

        if gaps:
            add_scan_log(scan_id, f"[GENERATOR] [STAGE 5] Auditor identified {len(gaps)} uncovered features. Generating gap test cases...")
            for gap in gaps:
                m_name = gap["module"]
                f_name = gap["feature"]
                
                gap_prompt = f"""
                Generate test cases for feature '{f_name}' in module '{m_name}'.
                Return ONLY JSON list.
                """
                try:
                    gap_res = query_llm(model, gap_prompt, scan_id, f"STAGE 5: Auditor Gap - {f_name}")
                    gap_cases = json.loads(gap_res)
                    for gc in gap_cases:
                        tc_count += 1
                        gc["test_case_id"] = f"TC-{m_name.upper().replace(' ', '_')[:6]}-GAP-{tc_count}"
                        if "page" not in gc:
                            gc["page"] = "General"
                        generated_test_cases.append(gc)
                except Exception as e:
                    add_scan_log(scan_id, f"[WARNING] Gap generation failed: {e}")

        # --- STAGE 6: Quality Gate rewriting pass ---
        raw_test_cases = run_quality_gate(model, generated_test_cases, scan_id, scope_ctx=scope_ctx)

        # Deterministic safety net: the Quality Gate's DOM/CSS rejection rule is LLM-driven and
        # can miss cases, so also apply the rule-based structural-noise guard here directly.
        noise_dropped = [tc for tc in raw_test_cases if is_structural_noise_scenario(tc.get("scenario", ""))]
        if noise_dropped:
            log_pipeline(scan_id, scope_ctx, "QUALITY-GATE", f"Dropped {len(noise_dropped)} DOM/CSS structural-noise scenario(s) that survived the Quality Gate.")
        raw_test_cases = [tc for tc in raw_test_cases if not is_structural_noise_scenario(tc.get("scenario", ""))]

        # --- STAGE 7: Python Telemetry Classification Overwrite ---
        log_pipeline(scan_id, scope_ctx, "GENERATOR", "Running Immutable Python Classification Overrides...")
        final_test_cases = []
        for tc in raw_test_cases:
            m_val = str(tc.get("module", "")).strip().lower()
            f_val = str(tc.get("feature", "")).strip().lower()
            
            # Map module name variants if needed
            matched_status = "INFERRED_SCENARIO"
            if total_mcp_actions > 0:
                for (reg_mod, reg_feat), status in feature_registry.items():
                    if reg_feat == f_val and (reg_mod in m_val or m_val in reg_mod):
                        matched_status = status
                        break
            else:
                matched_status = "STRUCTURAL_EVIDENCE" if ("submit" in f_val or "form" in f_val or "button" in f_val) else "INFERRED_SCENARIO"
                    
            tc["observed_behavior"] = matched_status
            tc["source_classification"] = matched_status
            if matched_status == "OBSERVED_BEHAVIOR":
                tc["evidence_reference"] = f"EVID-{tc.get('test_case_id')}"
            else:
                tc["evidence_reference"] = None

            # Best-effort automation metadata: attach a locator hint if this test case's
            # feature matches a known discovered feature. Kept as a separate field, never
            # exposed as the scenario text itself.
            automation_selector = ""
            for (reg_mod, reg_feat), hint in locator_registry.items():
                if reg_feat == f_val and (reg_mod in m_val or m_val in reg_mod):
                    automation_selector = hint
                    break
            tc["automation_selector"] = automation_selector

            # Which engine actually drove this feature (MCP / PLAYWRIGHT_CORE / NONE) - never
            # label Playwright Core execution as MCP or vice versa.
            execution_engine = "NONE"
            for (reg_mod, reg_feat), engine in engine_registry.items():
                if reg_feat == f_val and (reg_mod in m_val or m_val in reg_mod):
                    execution_engine = engine
                    break
            tc["execution_engine"] = execution_engine

            final_test_cases.append(tc)

        # Strict Excel and context boundary check for MODULE scope
        if scope_ctx and scope_ctx.scan_scope in ("module", "MODULE", "MODULE_WISE"):
            filtered_cases = []
            for tc in final_test_cases:
                tc_mod = str(tc.get("module", "")).strip().lower()
                tc_sub = str(tc.get("submodule", "")).strip().lower()
                
                parent_match = scope_ctx.parent_module.strip().lower()
                selected_match = scope_ctx.selected_module.strip().lower()
                
                is_parent_match = (parent_match in tc_mod or tc_mod in parent_match or parent_match in tc_sub or tc_sub in parent_match)
                is_selected_match = (selected_match in tc_mod or tc_mod in selected_match or selected_match in tc_sub or tc_sub in selected_match)
                
                if is_parent_match or is_selected_match:
                    tc["module"] = scope_ctx.parent_module
                    tc["submodule"] = scope_ctx.selected_module
                    filtered_cases.append(tc)
                else:
                    log_pipeline(scan_id, scope_ctx, "EXCEL-SCOPE-VALIDATION", f"Dropping test case out of module scope: {tc.get('test_case_id')} | module={tc.get('module')} | scenario={tc.get('scenario')}")
            
            final_test_cases = filtered_cases
            log_pipeline(scan_id, scope_ctx, "EXCEL-SCOPE-VALIDATION", f"Retained {len(final_test_cases)} test cases strictly within module scope.")

        # --- STAGE 8: Validation Metrics & Matrix Log ---
        total_modules = len(set(tc.get("module", "") for tc in final_test_cases))
        total_pages = len(set(tc.get("page", "") for tc in final_test_cases))
        total_features = len(set((tc.get("module", ""), tc.get("feature", "")) for tc in final_test_cases))
        
        observed_count = sum(1 for tc in final_test_cases if tc.get("observed_behavior") == "OBSERVED_BEHAVIOR")
        structural_count = sum(1 for tc in final_test_cases if tc.get("observed_behavior") == "STRUCTURAL_EVIDENCE")
        inferred_count = sum(1 for tc in final_test_cases if tc.get("observed_behavior") == "INFERRED_SCENARIO")
        skipped_count = sum(1 for tc in final_test_cases if tc.get("observed_behavior") == "SKIPPED_SCENARIO")
        
        coverage_pct = int((observed_count / max(1, observed_count + structural_count + inferred_count + skipped_count)) * 100) if final_test_cases else 100
        
        log_pipeline(scan_id, scope_ctx, "GENERATOR", "==================================================")
        log_pipeline(scan_id, scope_ctx, "GENERATOR", "               VALIDATION METRICS                 ")
        log_pipeline(scan_id, scope_ctx, "GENERATOR", "==================================================")
        log_pipeline(scan_id, scope_ctx, "GENERATOR", f" - Total Modules: {total_modules}")
        log_pipeline(scan_id, scope_ctx, "GENERATOR", f" - Total Pages: {total_pages}")
        log_pipeline(scan_id, scope_ctx, "GENERATOR", f" - Total Features: {total_features}")
        log_pipeline(scan_id, scope_ctx, "GENERATOR", f" - Observed Behaviors: {observed_count}")
        log_pipeline(scan_id, scope_ctx, "GENERATOR", f" - Structurally Discovered Behaviors: {structural_count}")
        log_pipeline(scan_id, scope_ctx, "GENERATOR", f" - Inferred Behaviors: {inferred_count}")
        log_pipeline(scan_id, scope_ctx, "GENERATOR", f" - Skipped Behaviors: {skipped_count}")
        log_pipeline(scan_id, scope_ctx, "GENERATOR", f" - Coverage %: {coverage_pct}%")
        log_pipeline(scan_id, scope_ctx, "GENERATOR", f" - Generated Test Cases: {len(final_test_cases)}")
        log_pipeline(scan_id, scope_ctx, "GENERATOR", "==================================================")

        # Print Coverage Matrix
        log_pipeline(scan_id, scope_ctx, "GENERATOR", "=======================================================================================")
        log_pipeline(scan_id, scope_ctx, "GENERATOR", "                           FEATURE COVERAGE MATRIX                                     ")
        log_pipeline(scan_id, scope_ctx, "GENERATOR", "=======================================================================================")
        log_pipeline(scan_id, scope_ctx, "GENERATOR", " Module | Feature | Observed | Structural | Inferred | Skipped | Coverage % ")
        log_pipeline(scan_id, scope_ctx, "GENERATOR", "---------------------------------------------------------------------------------------")
        unique_features = set((tc.get("module", ""), tc.get("feature", "")) for tc in final_test_cases)
        for mod, feat in sorted(unique_features):
            feat_cases = [tc for tc in final_test_cases if tc.get("module") == mod and tc.get("feature") == feat]
            obs = sum(1 for tc in feat_cases if tc.get("observed_behavior") == "OBSERVED_BEHAVIOR")
            struct = sum(1 for tc in feat_cases if tc.get("observed_behavior") == "STRUCTURAL_EVIDENCE")
            inf = sum(1 for tc in feat_cases if tc.get("observed_behavior") == "INFERRED_SCENARIO")
            skip = sum(1 for tc in feat_cases if tc.get("observed_behavior") == "SKIPPED_SCENARIO")
            total = obs + struct + inf + skip
            feat_pct = int((obs / max(1, total)) * 100)
            log_pipeline(scan_id, scope_ctx, "GENERATOR", f" {mod} | {feat} | {obs} | {struct} | {inf} | {skip} | {feat_pct}% ")
        log_pipeline(scan_id, scope_ctx, "GENERATOR", "=======================================================================================")

        return final_test_cases
        
    except Exception as e:
        log_pipeline(scan_id, scope_ctx, "GENERATOR", f"[ERROR] Failed during AI reasoning: {e}. Falling back to mock generator.")
        return get_mock_test_cases(app_map, scope_ctx)

def export_to_excel(test_cases: list, filename: str) -> str:
    """Exports test cases JSON array to a sorted, styled, premium formatted Excel sheet with full telemetry mapping."""
    df = pd.DataFrame(test_cases)
    
    cols = [
        "module",
        "submodule",
        "page",
        "feature",
        "test_case_id",
        "scenario",
        "test_type",
        "priority",
        "preconditions",
        "test_data",
        "test_steps",
        "expected_result",
        "business_rule",
        "dependencies",
        "observed_behavior",
        "evidence_reference",
        "automation_candidate",
        "automation_selector",
        "execution_engine"
    ]
    for col in cols:
        if col not in df.columns:
            df[col] = ""
            
    df = df[cols]
    
    # Rename columns to Title Case for sheet display
    df.columns = [c.replace("_", " ").title() for c in cols]
    
    # Sort dataframe logically by Module and Feature
    df = df.sort_values(by=["Module", "Feature"])
    
    file_path = os.path.join(settings.DOWNLOADS_DIR, filename)
    
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Test Cases")
        
        workbook = writer.book
        worksheet = writer.sheets["Test Cases"]
        
        # Header style (indigo/blue theme)
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo 600
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        
        data_font = Font(name="Segoe UI", size=10, bold=False)
        center_align = Alignment(horizontal="center", vertical="top")
        left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'), # Light grey borders
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        
        # Priority pastel fills
        priority_colors = {
            "high": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),  # Light Red
            "medium": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),# Light Amber
            "low": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")   # Light Emerald
        }
        
        # Apply header formatting
        for col_idx in range(1, len(cols) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
        # Apply cell/row data formatting
        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(cols) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                
                col_name = cols[col_idx - 1]
                if col_name in ["test_case_id", "priority", "test_type", "automation_candidate", "observed_behavior", "evidence_reference", "execution_engine"]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
                    
                # Highlight priority
                if col_name == "priority":
                    val = str(cell.value or "").lower().strip()
                    if val in priority_colors:
                        cell.fill = priority_colors[val]
                        
        # Auto-adjust columns widths
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                lines = val_str.split('\n')
                line_max = max(len(l) for l in lines) if lines else 0
                if line_max > max_len:
                    max_len = line_max
            worksheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)
            
    return file_path
